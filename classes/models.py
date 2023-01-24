from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image


class LeitorDeAcoes:
    def __init__(self, caminho_arquivo: str = ""):
        self.caminho_arquivo = caminho_arquivo
        self.dados = []

    def processa_arquivo(self, acao: str = ""):
        with open(f'{self.caminho_arquivo}{acao}.txt', 'r') as arquivo_cotacao:
            linhas = arquivo_cotacao.readlines()
            self.dados = [linha.removesuffix('\n').split(';') for linha in linhas]


class PropriedadeSerieGrafico:
    def __init__(self, grossura: int, cor_preenchimento: str):
        self.grossura = grossura
        self.cor_preenchimento = cor_preenchimento


class GerenciadorPlanilha:
    def __init__(self):
        self.workbook = Workbook()
        self.planilha_ativa = None

    def adicionar_planilha(self, titulo_planilha: str = ""):
        nova_planilha = self.workbook.create_sheet(titulo_planilha)
        self.workbook.active = nova_planilha
        self.planilha_ativa = nova_planilha

        return nova_planilha

    def adiciona_linha(self, dados: list):
        self.planilha_ativa.append(dados)

    def atualiza_celula(self, celula: str, dado):
        self.planilha_ativa[celula] = dado

    def mescla_celulas(self, celula_inicio: str, celula_fim: str):
        self.planilha_ativa.merge_cells(f'{celula_inicio}:{celula_fim}')

    def aplica_estilos(self, celula: str, estilos: list[tuple]):
        for estilo in estilos:
            setattr(self.planilha_ativa[celula], estilo[0], estilo[1])

    def adiciona_grafico_linha(self, celula: str, comprimento: float, altura: float, titulo: str,
                               titulo_eixo_x: str, titulo_eixo_y: str,
                               referencia_eixo_x: Reference, referencia_eixo_y: Reference,
                               propriedades_grafico: list
                               ):
        grafico = LineChart()
        grafico.width = comprimento
        grafico.height = altura
        grafico.title = titulo
        grafico.x_axis.title = titulo_eixo_x
        grafico.y_axis.title = titulo_eixo_y

        grafico.add_data(referencia_eixo_x)
        grafico.set_categories(referencia_eixo_y)

        for serie, propriedade in zip(grafico.series, propriedades_grafico):
            serie.graphicalProperties.line.width = propriedade.grossura
            serie.graphicalProperties.line.solidFill = propriedade.cor_preenchimento

        self.planilha_ativa.add_chart(grafico, celula)

    def adiciona_grafico_linha_2(self, celula: str, referencia_eixo_x: Reference,
                                 referencia_eixo_y: Reference, propriedades_grafico: list[tuple],
                                 propriedades_series: list[tuple[int, str, str | int]]):
        grafico = LineChart()
        grafico.add_data(referencia_eixo_x)
        grafico.set_categories(referencia_eixo_y)
        for propriedade in propriedades_grafico:
            setattr(grafico, propriedade[0], propriedade[1])

        for propriedade in propriedades_series:
            match propriedade:
                case index, 'width', valor:
                    grafico.series[index].graphicalProperties.line.width = valor
                case index, 'solidFill', valor:
                    grafico.series[index].graphicalProperties.line.solidFill = valor
                case _, _, _:
                    pass

        self.planilha_ativa.add_chart(grafico, celula)

    def adiciona_imagem(self, celula: str, caminho_imagem: str):
        imagem = Image(caminho_imagem)
        self.planilha_ativa.add_image(imagem, celula)

    def salva_arquivo(self, caminho_arquivo):
        self.workbook.save(caminho_arquivo)
